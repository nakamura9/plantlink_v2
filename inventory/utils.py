import csv
from io import StringIO, BytesIO
from inventory.models import Machine, Section, SubAssembly, SubUnit, Component
import openpyxl


def parse_file(file, filename):
    if "xls" in filename or "xlx" in filename:
        print("loading")
        content = BytesIO(file.read())
        wb = openpyxl.load_workbook(filename=content)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        rows = []
        current = 1
        while True:
            A = ws[f"A{current}"].value
            if not A:
                break
            B = ws[f"B{current}"].value
            rows.append((A, B))
            current += 1
        
    elif "csv" in filename:
        content = StringIO(file.read().decode("utf-8", "ignore"))
        rows = csv.reader(content)
    else:
        raise Exception("Invalid file type")

    for i, row in enumerate(rows):
        try:
            num = int(row[0])
        except:
            print(f"error, missing data in row {i}")
            continue

        if len(str(num)) % 2 != 0:
            id_string =  "0" + str(num)
        else:
            id_string = str(num)
        name = row[1]
        print(name, id_string)
        try: 
            if len(id_string) == 4:     
                machine = Machine.objects.get(pk=id_string[:2])
                print('section created')
                Section.objects.create(
                    unique_id=id_string,
                    section_name=name,
                    machine=machine
                )
                
            elif len(id_string) == 6:
                #Subt Unit
                machine = Machine.objects.get(pk=id_string[:2])
                section = Section.objects.get(pk=id_string[:4])
                SubUnit.objects.create(
                    unique_id=id_string,
                    unit_name=name,
                    machine=machine,
                    section=section
                )

            elif len(id_string) == 8:
                #Sub assembly
                machine = Machine.objects.get(pk=id_string[:2])
                section = Section.objects.get(pk=id_string[:4])
                subunit = SubUnit.objects.get(pk=id_string[:6])
                SubAssembly.objects.create(
                    unique_id=id_string,
                    unit_name=name,
                    machine = machine,
                    section = section,
                    subunit = subunit
                )

            elif len(id_string) == 10:
                #Component
                machine = Machine.objects.get(pk=id_string[:2])
                section = Section.objects.get(pk=id_string[:4])
                subunit = SubUnit.objects.get(pk=id_string[:6])
                subassy = SubAssembly.objects.get(pk=id_string[:8])
                Component.objects.create(
                    unique_id=id_string,
                    component_name=name,
                    machine=machine,
                    section=section,
                    subunit=subunit,
                    subassembly=subassy
                )
        except Exception as e:
            print(e)
            